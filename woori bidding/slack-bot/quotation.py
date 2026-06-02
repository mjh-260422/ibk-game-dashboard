# quotation.py
import os
import re
import shutil
import subprocess
import sys
from copy import copy

import openpyxl
from openpyxl.cell.cell import MergedCell

from image_parser import BidInfo

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "윈큐브_견적서_template.xlsx")


def _copy_style(src, dst) -> None:
    if isinstance(dst, MergedCell):
        return
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _brand_name(product_name: str) -> str:
    m = re.search(r'\[(.+?)\]', product_name)
    return m.group(1) if m else product_name.split()[0]


def create_quotation_excel(bid: BidInfo, output_dir: str) -> str:
    safe = bid.event_name.replace("'", "").replace("/", "-").strip()
    filename = f"윈큐브 견적서_{safe}_{bid.bid_number}.xlsx"
    out_path = os.path.join(output_dir, filename)

    shutil.copy(TEMPLATE_PATH, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    prods = bid.products
    n = len(prods)
    template_n = 2  # 템플릿 기본 상품 행 수

    if n > template_n:
        extra = n - template_n
        ws.insert_rows(18, extra)
        for i in range(extra):
            nr = 18 + i
            for col in range(1, ws.max_column + 1):
                _copy_style(ws.cell(17, col), ws.cell(nr, col))
            ws.merge_cells(f"D{nr}:F{nr}")
    elif n < template_n:
        for er in range(16 + n, 16 + template_n):
            to_remove = [str(m) for m in ws.merged_cells.ranges if m.min_row <= er <= m.max_row]
            for m in to_remove:
                ws.unmerge_cells(m)
        ws.delete_rows(16 + n, template_n - n)

    for i, p in enumerate(prods):
        r = 16 + i
        ws.cell(r, 2).value = i + 1
        ws.cell(r, 3).value = _brand_name(p.name)
        ws.cell(r, 4).value = p.name
        ws.cell(r, 7).value = p.quantity
        ws.cell(r, 8).value = p.face_value
        ws.cell(r, 9).value = p.discount_rate / 100
        ws.cell(r, 10).value = f"=ROUND(H{r}-(H{r}*I{r}),0)"
        ws.cell(r, 11).value = f"=J{r}*G{r}"

    tr = 16 + n
    ws.cell(tr, 7).value = f"=SUM(G16:G{15+n})"
    ws.cell(tr, 10).value = f"=SUM(K16:K{15+n})"

    wb.save(out_path)
    return out_path


def convert_to_pdf(xlsx_path: str) -> str:
    pdf_path = xlsx_path.replace(".xlsx", ".pdf")
    out_dir = os.path.dirname(xlsx_path)

    if sys.platform == "win32":
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
    else:
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice PDF 변환 실패: {result.stderr}")

    return pdf_path
