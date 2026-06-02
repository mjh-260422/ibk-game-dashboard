import os, tempfile
from image_parser import BidInfo, ProductInfo
from quotation import create_quotation_excel, TEMPLATE_PATH

def _make_bid_single():
    return BidInfo(bid_number="1771", event_name="법인사업자 기업카드 무실적",
                   manager="김소연", is_pin_delivery=False,
                   products=[ProductInfo(name="베스킨라빈스 싱글레귤러 2개세트",
                                         face_value=7800, quantity=826, discount_rate=20.0)])

def _make_bid_multi():
    return BidInfo(bid_number="1772", event_name="'26년 4월 쿠팡 경품",
                   manager="박찬규", is_pin_delivery=False,
                   products=[
                       ProductInfo(name="[BHC] 뿌링클 세트", face_value=23500, quantity=12, discount_rate=8.0),
                       ProductInfo(name="[스타벅스] 아이스 아메리카노 Tall", face_value=4700, quantity=110, discount_rate=25.0),
                   ])

def test_template_exists():
    assert os.path.exists(TEMPLATE_PATH), f"템플릿 파일 없음: {TEMPLATE_PATH}"

def test_create_quotation_single_product():
    import openpyxl
    bid = _make_bid_single()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = create_quotation_excel(bid, tmpdir)
        assert os.path.exists(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(16, 3).value == "베스킨라빈스"  # C: 구분
        assert ws.cell(16, 7).value == 826              # G: 수량
        assert ws.cell(16, 8).value == 7800             # H: 정상가
        assert abs(ws.cell(16, 9).value - 0.20) < 1e-9  # I: 할인율

def test_create_quotation_multi_product():
    import openpyxl
    bid = _make_bid_multi()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = create_quotation_excel(bid, tmpdir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(16, 7).value == 12   # 첫 상품 수량
        assert ws.cell(17, 7).value == 110  # 두번째 상품 수량
