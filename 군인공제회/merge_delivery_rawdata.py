import os, re, sys
from bs4 import BeautifulSoup
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'c:\Users\jihye\tutorial\군인공제회\202508-202603 발송별 로우데이터'
OUTPUT = r'c:\Users\jihye\tutorial\군인공제회\202508-202604_발송별_병합.xlsx'

# ── 컬럼 정의 ──────────────────────────────────────────
COLS = ['월','담당자','발송일','등록일','광고주','캠페인번호','발송번호',
        '캠페인명','발송명','상품공급사','교환처','판매상품명','판매가',
        '상품공급수수료(%)','상품공급가','제공할인율(%)','제공가','공급사정산기준',
        '발송량','교환량','교환율(%)','업체제공가','상품대금',
        '발송비','보전받는금액','발송수익','판매수익','미사용수익',
        '총수익','수익율(%)','거래금액','판매가x수량','미교환율(%)','기간만료일']

def num(v):
    try: return float(str(v).replace(',','').strip())
    except: return None

def parse_file(fpath, month_label):
    with open(fpath, encoding='euc-kr', errors='replace') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')
    rows = soup.find_all('tr')
    records = []
    for row in rows[2:]:  # 헤더 2행 제외
        cells = [td.get_text(strip=True) for td in row.find_all('td')]
        if len(cells) >= 33:
            records.append([month_label] + cells[:33])
    return records

# ── 전체 파싱 ──────────────────────────────────────────
files = sorted([f for f in os.listdir(FOLDER) if f.endswith('.xls')])
all_records = []
for fname in files:
    m = re.search(r'(\d{4})(\d{2})\d{2}~', fname)
    month = f"{m.group(1)}-{m.group(2)}" if m else fname
    recs = parse_file(os.path.join(FOLDER, fname), month)
    all_records.extend(recs)
    print(f'{month}: {len(recs)}건')

print(f'총 {len(all_records)}건')

# ── 스타일 ──────────────────────────────────────────
wb = openpyxl.Workbook()
thin = Side(style='thin', color='D0D0D0')
BD   = Border(left=thin, right=thin, top=thin, bottom=thin)
H1   = PatternFill('solid', fgColor='1F4E79')
H2   = PatternFill('solid', fgColor='2E75B6')
H3   = PatternFill('solid', fgColor='BDD7EE')
ALT  = PatternFill('solid', fgColor='F5F9FF')
SUM  = PatternFill('solid', fgColor='DEEAF1')
NOTE = PatternFill('solid', fgColor='FFFACD')
POS  = PatternFill('solid', fgColor='E2EFDA')
NEG  = PatternFill('solid', fgColor='FCE4D6')
WF   = Font(color='FFFFFF', bold=True, size=9)
BF   = Font(bold=True, size=9)
NF   = Font(size=9)
GF   = Font(color='375623', bold=True, size=9)
RF   = Font(color='C00000', bold=True, size=9)
CA   = Alignment(horizontal='center', vertical='center', wrap_text=True)
RA   = Alignment(horizontal='right',  vertical='center')
LA   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def cs(ws, r, c, v, fill=None, font=NF, align=RA, fmt=None):
    cell = ws.cell(r, c, v)
    if fill: cell.fill = fill
    cell.font  = font
    cell.alignment = align
    if fmt: cell.number_format = fmt
    cell.border = BD
    return cell

# ══════════════════════════════════════════════════════
# 시트 1: 전체 병합
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '발송별 전체 병합'

# 컬럼 너비
widths = [9,9,11,11,20,14,14,40,40,14,16,26,8,9,8,9,8,9,7,7,8,10,12,8,10,10,10,10,8,8,12,12,8,12]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 헤더
for ci, h in enumerate(COLS, 1):
    cs(ws1, 1, ci, h, fill=H1, font=WF, align=CA)
ws1.row_dimensions[1].height = 30
ws1.freeze_panes = 'C2'

# 데이터
num_cols = {13,14,15,16,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32}
pct_cols = {14,16,20,28,30,32}   # % 표시 컬럼 (인덱스)
money_cols = {13,15,16,18,19,21,22,23,24,25,26,27,29,30}

for ri, rec in enumerate(all_records, 2):
    fill = ALT if ri % 2 == 0 else None
    for ci, v in enumerate(rec, 1):
        col_i = ci  # 1-based
        if col_i in num_cols:
            n = num(v)
            fmt = '0.00' if col_i in pct_cols else '#,##0'
            cs(ws1, ri, ci, n, fill=fill, fmt=fmt)
        else:
            cs(ws1, ri, ci, v, fill=fill, align=CA if col_i == 1 else LA)

# ══════════════════════════════════════════════════════
# 시트 2: 인사이트
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet('인사이트')
ws2.column_dimensions['A'].width = 28
for c in 'BCDEFGHIJKLM':
    ws2.column_dimensions[c].width = 14

months  = sorted(set(r[0] for r in all_records))
MONTH_N = len(months)

# 권종 축약
def short(name):
    m2 = re.search(r'(\d+)만원', str(name))
    if m2:
        w = int(m2.group(1))
        kind = '편의점' if '편의점' in str(name) or 'CU' in str(name) or 'GS' in str(name) or '이마트24' in str(name) or '세븐' in str(name) else '백화점'
        return f'{kind} {w}만원'
    return str(name)[:20]

# 레코드를 분석용 dict로 변환
records_d = []
for rec in all_records:
    try:
        if '테스트' in rec[7]:  # 캠페인명 테스트 제외
            continue
        records_d.append({
            '월':       rec[0],
            '발송일':   rec[2],
            '캠페인명': rec[7],
            '발송명':   rec[8],
            '상품명':   rec[11],
            '권종':     short(rec[11]),
            '판매가':   num(rec[12]),
            '수수료':   num(rec[13]),
            '공급가':   num(rec[14]),
            '할인율':   num(rec[15]),
            '제공가':   num(rec[16]),
            '발송량':   num(rec[18]),
            '교환량':   num(rec[19]),
            '교환율':   num(rec[20]),
            '상품대금': num(rec[22]),
            '총수익':   num(rec[28]),
            '수익율':   num(rec[29]),
            '거래금액': num(rec[30]),
            '판매금액': num(rec[31]),
            '미교환율': num(rec[32]),
            '만료일':   rec[33],
        })
    except:
        pass

r = 1

# ── 섹션 헬퍼 ──
def sec(ws, r, title, ncols=len(months)+2):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1, title)
    c.fill = H2; c.font = WF; c.alignment = CA; c.border = BD
    ws.row_dimensions[r].height = 20
    return r + 1

def hrow(ws, r, labels, fills=None):
    for ci, lab in enumerate(labels, 1):
        f = (fills[ci-1] if fills else H3)
        cs(ws, r, ci, lab, fill=f, font=BF, align=CA)
    return r + 1

# ═══ 블록 1: 기본 정보 요약 ═══════════════════════════
r = sec(ws2, r, '■ 기본 정보 요약 (전체 계약 기간: 2025.08 ~ 2026.04)')

labels = ['항목', '내용']
r = hrow(ws2, r, labels)
total_거래 = sum(d['거래금액'] or 0 for d in records_d)
total_발송 = sum(d['발송량']  or 0 for d in records_d)
total_교환 = sum(d['교환량']  or 0 for d in records_d)
total_수익 = sum(d['총수익']  or 0 for d in records_d)
n_months   = len(months)
담당자목록 = list(dict.fromkeys(r2[1] for r2 in all_records))

basics = [
    ('계약 기간',         f"{months[0]} ~ {months[-1]} ({n_months}개월)"),
    ('총 발송 레코드 수', f"{len(records_d)}건 (월평균 {len(records_d)/n_months:.1f}건)"),
    ('총 거래금액 합산',  f"{total_거래:,.0f}원 (월평균 {total_거래/n_months:,.0f}원)"),
    ('총 발송량',         f"{total_발송:,.0f}매 (월평균 {total_발송/n_months:,.0f}매)"),
    ('총 교환량',         f"{total_교환:,.0f}매"),
    ('전체 교환율',       f"{total_교환/total_발송*100:.1f}%" if total_발송 else '-'),
    ('전체 미교환율',     f"{(total_발송-total_교환)/total_발송*100:.1f}%" if total_발송 else '-'),
    ('총수익 합산',       f"{total_수익:,.0f}원"),
    ('담당자 변경 이력',  ' → '.join(담당자목록)),
]
for item, val in basics:
    cs(ws2, r, 1, item, fill=ALT if r%2==0 else None, font=BF, align=LA)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(months)+2)
    cs(ws2, r, 2, val, fill=ALT if r%2==0 else None, font=NF, align=LA)
    r += 1

r += 1

# ═══ 블록 2: 월별 제공할인율 (권종별) ════════════════
r = sec(ws2, r, '■ 월별 제공할인율(%) — 군인공제회에 제공한 실제 할인율', ncols=len(months)+2)
r = hrow(ws2, r, ['권종/발송명'] + months + ['비고'])

# (권종, 월) 기준으로 직접 집계 — 발송명 30자 truncation 키 충돌 방지
discount_map2 = defaultdict(lambda: defaultdict(set))
for d in records_d:
    if d['할인율'] is not None:
        discount_map2[d['권종']][d['월']].add(d['할인율'])

for prod in sorted(discount_map2.keys()):
    mdict = discount_map2[prod]
    all_rates = [list(mdict.get(mo, {None}))[0] for mo in months]
    changed = len(set(v for v in all_rates if v is not None)) > 1
    note = '★할인율 변경 있음' if changed else ''

    fill = NOTE if changed else None
    fn   = BF if changed else NF
    cs(ws2, r, 1, prod, fill=fill, font=fn, align=LA)
    for ci, (mo, rate) in enumerate(zip(months, all_rates), 2):
        v = f"{rate:.2f}%" if rate is not None else '-'
        cs(ws2, r, ci, v, fill=fill, font=fn, align=CA)
    cs(ws2, r, len(months)+2, note, fill=fill, font=RF if changed else NF, align=LA)
    r += 1

r += 1

# ═══ 블록 3: 월별 거래금액 ══════════════════════════
r = sec(ws2, r, '■ 월별 거래금액 (원)', ncols=len(months)+2)
r = hrow(ws2, r, ['권종'] + months + ['합계'])

권종_거래 = defaultdict(lambda: defaultdict(float))
for d in records_d:
    if d['거래금액']:
        권종_거래[d['권종']][d['월']] += d['거래금액']

all_권종 = sorted(권종_거래.keys())
월별합계 = defaultdict(float)

for i, prod in enumerate(all_권종):
    fill = ALT if i%2==0 else None
    cs(ws2, r, 1, prod, fill=fill, font=NF, align=LA)
    row_sum = 0
    for ci, mo in enumerate(months, 2):
        v = 권종_거래[prod].get(mo, 0)
        row_sum += v
        월별합계[mo] += v
        cs(ws2, r, ci, v if v else None, fill=fill, fmt='#,##0')
    cs(ws2, r, len(months)+2, row_sum, fill=fill, font=BF, fmt='#,##0')
    r += 1

# 합계 행
cs(ws2, r, 1, '합계', fill=SUM, font=BF, align=CA)
grand = 0
for ci, mo in enumerate(months, 2):
    v = 월별합계[mo]
    grand += v
    cs(ws2, r, ci, v, fill=SUM, font=BF, fmt='#,##0')
cs(ws2, r, len(months)+2, grand, fill=SUM, font=BF, fmt='#,##0')
r += 2

# ═══ 블록 4: 월별 미교환율 ══════════════════════════
r = sec(ws2, r, '■ 월별 미교환율(%) — 낙전 수익 핵심 지표', ncols=len(months)+2)
r = hrow(ws2, r, ['권종'] + months + ['평균'])

권종_미교환 = defaultdict(lambda: defaultdict(list))
for d in records_d:
    if d['미교환율'] is not None and d['발송량'] and d['발송량'] > 0:
        권종_미교환[d['권종']][d['월']].append((d['미교환율'], d['발송량']))

for i, prod in enumerate(all_권종):
    fill = ALT if i%2==0 else None
    cs(ws2, r, 1, prod, fill=fill, font=NF, align=LA)
    all_vals = []
    for ci, mo in enumerate(months, 2):
        entries = 권종_미교환[prod].get(mo, [])
        if entries:
            w_sum = sum(rate * qty for rate, qty in entries)
            q_sum = sum(qty for _, qty in entries)
            avg = w_sum / q_sum if q_sum else 0
            all_vals.append(avg)
            fill2 = NOTE if avg >= 40 else (POS if avg >= 20 else fill)
            cs(ws2, r, ci, avg, fill=fill2, fmt='0.0"%"')
        else:
            cs(ws2, r, ci, None, fill=fill)
    avg_all = sum(all_vals)/len(all_vals) if all_vals else 0
    cs(ws2, r, len(months)+2, avg_all, fill=NOTE if avg_all>=40 else SUM, font=BF, fmt='0.0"%"')
    r += 1

r += 1

# ═══ 블록 5: 공급수수료 변화 ══════════════════════
r = sec(ws2, r, '■ 상품공급 수수료율(%) 변화 — 브랜드 수수료 (낙찰 전후 변경 여부)', ncols=len(months)+2)
r = hrow(ws2, r, ['권종'] + months + ['비고'])

권종_수수료 = defaultdict(lambda: defaultdict(set))
for d in records_d:
    if d['수수료'] is not None:
        권종_수수료[d['권종']][d['월']].add(d['수수료'])

for i, prod in enumerate(all_권종):
    fill = ALT if i%2==0 else None
    cs(ws2, r, 1, prod, fill=fill, font=NF, align=LA)
    all_rates = []
    for ci, mo in enumerate(months, 2):
        vals = 권종_수수료[prod].get(mo, set())
        v = list(vals)[0] if len(vals)==1 else (f"복수:{vals}" if vals else None)
        cs(ws2, r, ci, f"{v:.2f}%" if isinstance(v, float) else (v or '-'), fill=fill, align=CA)
        if isinstance(v, float): all_rates.append(v)
    changed = len(set(all_rates)) > 1
    cs(ws2, r, len(months)+2, '★변경있음' if changed else '일정', fill=NOTE if changed else fill, font=RF if changed else NF, align=CA)
    r += 1

r += 1

# ═══ 블록 6: 총수익 및 수익율 ══════════════════════
r = sec(ws2, r, '■ 월별 총수익 합산 및 평균 수익율(%)', ncols=len(months)+2)
r = hrow(ws2, r, ['구분'] + months + ['합계/평균'])

월별_수익 = defaultdict(float)
월별_수익율 = defaultdict(list)
월별_거래합 = defaultdict(float)
for d in records_d:
    if d['총수익']:  월별_수익[d['월']] += d['총수익']
    if d['수익율']: 월별_수익율[d['월']].append(d['수익율'])
    if d['거래금액']: 월별_거래합[d['월']] += d['거래금액']

cs(ws2, r, 1, '총수익(원)', fill=ALT, font=BF, align=LA)
grand_수익 = 0
for ci, mo in enumerate(months, 2):
    v = 월별_수익[mo]
    grand_수익 += v
    f = POS if v >= 0 else NEG
    cs(ws2, r, ci, v, fill=f, font=GF if v>=0 else RF, fmt='#,##0')
cs(ws2, r, len(months)+2, grand_수익, fill=POS if grand_수익>=0 else NEG, font=GF if grand_수익>=0 else RF, fmt='#,##0')
r += 1

cs(ws2, r, 1, '평균 수익율(%)', fill=None, font=BF, align=LA)
avg_all_수익율 = []
for ci, mo in enumerate(months, 2):
    rates = 월별_수익율[mo]
    avg = sum(rates)/len(rates) if rates else 0
    avg_all_수익율.append(avg)
    f = POS if avg >= 0 else NEG
    cs(ws2, r, ci, avg, fill=f, font=GF if avg>=0 else RF, fmt='0.00%' if avg > 1 else '0.0000')
grand_avg = sum(avg_all_수익율)/len(avg_all_수익율) if avg_all_수익율 else 0
cs(ws2, r, len(months)+2, grand_avg, fill=POS if grand_avg>=0 else NEG, font=GF if grand_avg>=0 else RF, fmt='0.00%' if grand_avg > 1 else '0.0000')
r += 2

# ═══ 블록 7: 핵심 인사이트 ══════════════════════════
r = sec(ws2, r, '■ 핵심 인사이트 (26년 입찰 활용 포인트)', ncols=len(months)+2)

def ins_row(ws, r, txt, fill=None, bold=False):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(months)+2)
    c = ws.cell(r, 1, txt)
    c.fill = fill or PatternFill('solid', fgColor='F5F9FF')
    c.font = BF if bold else NF
    c.alignment = LA
    c.border = BD
    ws.row_dimensions[r].height = 18

# 할인율 변화 계산
할인율_by_prod_month = defaultdict(dict)
for d in records_d:
    if d['할인율'] is not None:
        할인율_by_prod_month[d['권종']][d['월']] = d['할인율']

ins_row(ws2, r, '【할인율 변화】', bold=True); r+=1
for prod, mdict in sorted(할인율_by_prod_month.items()):
    rates_str = ' / '.join(f"{mo}:{v:.2f}%" for mo, v in sorted(mdict.items()))
    ins_row(ws2, r, f"  {prod}: {rates_str}"); r+=1

ins_row(ws2, r, ''); r+=1
ins_row(ws2, r, '【총 발송 현황】', bold=True); r+=1
ins_row(ws2, r, f"  총 거래금액: {total_거래:,.0f}원 / 총 발송량: {total_발송:,.0f}매 / 전체 미교환율: {(total_발송-total_교환)/total_발송*100:.1f}%"); r+=1
ins_row(ws2, r, f"  총수익(발송별 기준): {total_수익:,.0f}원 (총 거래금액 대비 {total_수익/total_거래*100:.2f}%)"); r+=1

ins_row(ws2, r, ''); r+=1
ins_row(ws2, r, '【26년 입찰 포인트】', bold=True, fill=NOTE); r+=1
ins_row(ws2, r, '  ① 제공할인율 추이: 25년 이후 월별 할인율 변화 위 표 참고 → 26년 입찰 제시 할인율 결정 시 참고', fill=NOTE); r+=1
ins_row(ws2, r, '  ② 3만원권 미교환율이 지속적으로 높음 → 26년 3만원 수량 증가(10,484매) 낙전 수익 극대화 기회', fill=NOTE); r+=1
ins_row(ws2, r, '  ③ 공급수수료율 변화(2.7%→2.6%→2.5%) 확인 → 수익 계산 시 최신 수수료율 적용 필수', fill=NOTE); r+=1
ins_row(ws2, r, '  ④ 담당자: 25년 8월 pji0201 → 25년 9월~ 김명신 → 26년 4월~ 오유경 (담당자 변경 시 관계 재구축 필요)', fill=NOTE); r+=1

ws2.freeze_panes = 'B2'

wb.save(OUTPUT)
print(f'\n저장 완료: {OUTPUT}')
print(f'시트: {wb.sheetnames}')

# ══════════════════════════════════════════════════════
# 자체 검증
# ══════════════════════════════════════════════════════
print('\n=== 자체 검증 ===')
errors = []

# 1) 테스트 레코드 제외 확인
test_in = [d for d in records_d if '테스트' in d.get('캠페인명','')]
if test_in:
    errors.append(f'[오류] 테스트 레코드가 insights에 포함됨: {len(test_in)}건')
else:
    print(f'[OK] 테스트 레코드 제외 확인 (전체 {len(all_records)}건 → insights {len(records_d)}건)')

# 2) 총 거래금액 교차 검증 (Sheet1 합산 vs records_d 합산)
s1_total = sum(num(rec[30]) or 0 for rec in all_records if '테스트' not in rec[7])
s2_total = sum(d['거래금액'] or 0 for d in records_d)
if abs(s1_total - s2_total) > 1:
    errors.append(f'[오류] 거래금액 합산 불일치: Sheet1={s1_total:,.0f} vs insights={s2_total:,.0f}')
else:
    print(f'[OK] 거래금액 합산 일치: {s2_total:,.0f}원')

# 3) 2026-02 할인율 테이블에 35만원권 없는지 확인
if '백화점 35만원' in discount_map2:
    mo202602 = discount_map2['백화점 35만원'].get('2026-02')
    if mo202602:
        errors.append(f'[오류] 백화점 35만원 2026-02 할인율 여전히 존재: {mo202602}')
    else:
        print('[OK] 백화점 35만원 2026-02 없음 확인')
else:
    print('[OK] 백화점 35만원 권종 없음 (해당 권종 미발송)')

# 4) 권종별 할인율 표 — 모든 권종이 records_d 권종 목록과 일치하는지
all_권종_in_data = set(d['권종'] for d in records_d)
all_권종_in_table = set(discount_map2.keys())
missing = all_권종_in_data - all_권종_in_table
extra   = all_권종_in_table - all_권종_in_data
if missing:
    errors.append(f'[오류] 할인율 표에 누락된 권종: {missing}')
else:
    print(f'[OK] 할인율 표 권종 완전 일치: {sorted(all_권종_in_table)}')

# 5) 월별 데이터 건수 확인
월별건수 = defaultdict(int)
for d in records_d:
    월별건수[d['월']] += 1
월별건수_raw = defaultdict(int)
for rec in all_records:
    if '테스트' not in rec[7]:
        월별건수_raw[rec[0]] += 1
mismatch = [(mo, 월별건수[mo], 월별건수_raw[mo]) for mo in months if 월별건수[mo] != 월별건수_raw[mo]]
if mismatch:
    errors.append(f'[오류] 월별 건수 불일치: {mismatch}')
else:
    print(f'[OK] 월별 건수 일치: {dict(sorted(월별건수.items()))}')

# 6) 미교환율 범위 이상값
abnormal = [(d['월'], d['권종'], d['미교환율']) for d in records_d if d['미교환율'] is not None and (d['미교환율'] < 0 or d['미교환율'] > 100)]
if abnormal:
    errors.append(f'[오류] 미교환율 범위 이상: {abnormal}')
else:
    print('[OK] 미교환율 범위 정상 (0~100%)')

print()
if errors:
    for e in errors:
        print(e)
    print(f'\n총 {len(errors)}개 오류 발견')
else:
    print('✓ 전체 검증 통과 — 오류 없음')
