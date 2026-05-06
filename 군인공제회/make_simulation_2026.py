import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()

# ── 공통 스타일 ──────────────────────────────────────────
H_FILL   = PatternFill('solid', fgColor='1F4E79')  # 헤더 진파랑
H_FONT   = Font(color='FFFFFF', bold=True, size=9)
T_FILL   = PatternFill('solid', fgColor='2E75B6')  # 섹션 파랑
T_FONT   = Font(color='FFFFFF', bold=True, size=9)
POS_FILL = PatternFill('solid', fgColor='E2EFDA')  # 긍정 연두
NEG_FILL = PatternFill('solid', fgColor='FCE4D6')  # 부정 연주
SUM_FILL = PatternFill('solid', fgColor='BDD7EE')  # 합계 연파
GRAY     = PatternFill('solid', fgColor='F2F2F2')
NOTE_FILL= PatternFill('solid', fgColor='FFFACD')  # 메모 노랑
RED_FONT = Font(color='C00000', bold=True, size=9)
GRN_FONT = Font(color='375623', bold=True, size=9)
BLK_FONT = Font(bold=True, size=9)
NRM_FONT = Font(size=9)
thin = Side(style='thin', color='BFBFBF')
BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
C        = Alignment(horizontal='center', vertical='center', wrap_text=True)
R        = Alignment(horizontal='right',  vertical='center')
L        = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def cell_set(ws, r, c, val, fill=None, font=None, align=None, fmt=None, border=True):
    cell = ws.cell(r, c, val)
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if fmt:    cell.number_format = fmt
    if border: cell.border = BORDER
    return cell

def hdr(ws, r, c, val, fill=H_FILL):
    cell_set(ws, r, c, val, fill=fill, font=H_FONT, align=C)

def merge(ws, r1, c1, r2, c2, val='', fill=None, font=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell_set(ws, r1, c1, val, fill=fill, font=font, align=C)

# ══════════════════════════════════════════════════════════
# 2026 데이터
# ══════════════════════════════════════════════════════════
ITEMS = [
    # 상품명, 권종, 공급수수료, 발송량, 미교환율_긍정, 미교환율_부정
    ('편의점 1만원권',      10000,  0.020, 19562, 0.25, 0.15),
    ('신세계 상품권 3만원', 30000,  0.027,  10484, 0.45, 0.30),
    ('신세계 상품권 5만원', 50000,  0.027,   6296, 0.25, 0.15),
    ('신세계 상품권 7만원', 70000,  0.027,   4542, 0.22, 0.13),
    ('신세계 상품권 10만원',100000, 0.027,  3364, 0.22, 0.13),
    ('신세계 상품권 15만원',150000, 0.027,  2100, 0.18, 0.10),
    ('신세계 상품권 20만원',200000, 0.027,  1185, 0.17, 0.10),
    ('신세계 상품권 25만원',250000, 0.027,   642, 0.12, 0.07),
    ('신세계 상품권 30만원',300000, 0.027,   344, 0.08, 0.03),
    ('신세계 상품권 35만원',350000, 0.027,    90, 0.05, 0.00),
    ('신세계 상품권 40만원',400000, 0.027,     3, 0.05, 0.00),
]

DISCOUNT_RATES = [0.07, 0.09, 0.11]  # 시뮬레이션할 할인율

# ══════════════════════════════════════════════════════════
# 시트 1: 시뮬레이션 (긍정)
# ══════════════════════════════════════════════════════════
def make_sim_sheet(wb, title, scenario, note_text):
    ws = wb.create_sheet(title)
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 30
    col_w = [22, 8, 8, 8, 8, 14, 8, 14, 16, 14, 10, 14, 14, 10]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    미교환율_idx = 4 if scenario == 'pos' else 5  # ITEMS index

    r = 1
    # 노트 박스
    ws.merge_cells(f'N1:N{3+len(ITEMS)+1}')
    note = ws['N1']
    note.value = note_text
    note.fill = NOTE_FILL
    note.font = Font(size=8)
    note.alignment = L

    # 할인율별 블록
    for di, disc in enumerate(DISCOUNT_RATES):
        if di > 0:
            r += 2

        # 섹션 헤더
        merge(ws, r, 1, r, 13, f'▶ 할인율 {disc*100:.0f}% 시나리오', fill=T_FILL, font=T_FONT)
        r += 1

        # 컬럼 헤더
        hdrs = ['상품명', '권종', '공급\n수수료', '발송량', '거래금액',
                '할인율', '할인금액', '청구금액',
                '예상\n수수료수익', '예상\n미교환율', '예상\n미교환수익',
                '예상 수익', '예상 수익율']
        for ci, h in enumerate(hdrs, 1):
            hdr(ws, r, ci, h)
        r += 1

        total_거래 = total_할인 = total_청구 = total_수수료 = total_미교환 = total_수익 = 0

        for item in ITEMS:
            name, 권종, 수수료율, 발송량, 미교환율_pos, 미교환율_neg = item
            미교환율 = 미교환율_pos if scenario == 'pos' else 미교환율_neg
            교환율 = 1 - 미교환율

            거래금액    = 권종 * 발송량
            할인금액    = round(거래금액 * disc)
            청구금액    = 거래금액 - 할인금액
            수수료수익  = round(거래금액 * 교환율 * 수수료율)
            미교환수익  = round(거래금액 * 미교환율)
            수익        = 미교환수익 + 수수료수익 - 할인금액
            수익율      = 수익 / 거래금액 if 거래금액 else 0

            row_fill = POS_FILL if scenario == 'pos' else NEG_FILL
            vals = [name, 권종, 수수료율, 발송량, 거래금액,
                    disc, 할인금액, 청구금액, 수수료수익,
                    미교환율, 미교환수익, 수익, 수익율]
            fmts = [None,'#,##0','0.0%','#,##0','#,##0',
                    '0%','#,##0','#,##0','#,##0',
                    '0.0%','#,##0','#,##0','0.00%']
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                cell = cell_set(ws, r, ci, v, fill=row_fill, font=NRM_FONT, align=R if ci > 1 else L, fmt=f)

            total_거래   += 거래금액
            total_할인   += 할인금액
            total_청구   += 청구금액
            total_수수료 += 수수료수익
            total_미교환 += 미교환수익
            total_수익   += 수익
            r += 1

        # 합계 행
        total_수익율 = total_수익 / total_거래 if total_거래 else 0
        sum_vals = ['합계', '', '', '', total_거래, disc, total_할인, total_청구,
                    total_수수료, '', total_미교환, total_수익, total_수익율]
        sum_fmts = [None, None, None, None, '#,##0', '0%', '#,##0', '#,##0',
                    '#,##0', None, '#,##0', '#,##0', '0.00%']
        for ci, (v, f) in enumerate(zip(sum_vals, sum_fmts), 1):
            fn = GRN_FONT if total_수익 > 0 else RED_FONT
            cell_set(ws, r, ci, v, fill=SUM_FILL, font=BLK_FONT if ci == 1 else fn, align=R if ci > 1 else L, fmt=f)
        r += 1

    ws.freeze_panes = 'A4'
    return ws

note_pos = ('【긍정 시나리오 근거】\n'
            '■ 25년 로우데이터 실제 미교환율\n'
            '  3만원: avg 50.3% → 45% 적용\n'
            '  5만원: avg 28.7% → 25% 적용\n'
            '  7만원: avg 26.4% → 22% 적용\n'
            ' 10만원: avg 26.5% → 22% 적용\n'
            ' 15만원: avg 20.7% → 18% 적용\n'
            ' 20만원: avg 19.9% → 17% 적용\n'
            ' 25만원: avg 15.0% → 12% 적용\n'
            ' 30만원: avg  9.6% →  8% 적용\n'
            ' 35만원: 데이터 부족 →  5% 가정\n'
            ' 40만원: 데이터 없음 →  5% 가정\n'
            ' 편의점: 직접 데이터 없음 → 25% 가정\n\n'
            '■ 24년 낙찰: 에스티엠 총할인율 10.87%\n'
            '■ 25년 예상: 총할인율 11.18%\n'
            '■ 26년 예산: 23억원\n'
            '■ 26년 기초금액 추정: 약 22.87억')

note_neg = ('【부정 시나리오 근거】\n'
            '■ 미교환율 보수적 하한 적용\n'
            '  3만원: 30% (실제 최솟값 38.9%보다 낮게)\n'
            '  5만원: 15%\n'
            '  7만원: 13%\n'
            ' 10만원: 13%\n'
            ' 15만원: 10%\n'
            ' 20만원: 10%\n'
            ' 25만원:  7%\n'
            ' 30만원:  3%\n'
            ' 35만원:  0%\n'
            ' 40만원:  0%\n'
            ' 편의점:  15%\n\n'
            '■ 미교환율 예측 빗나갈 경우 최악의 수익 시나리오')

make_sim_sheet(wb, '시뮬레이션 (긍정)', 'pos', note_pos)
make_sim_sheet(wb, '시뮬레이션 (부정)', 'neg', note_neg)

# ══════════════════════════════════════════════════════════
# 시트 3: 입찰 전략
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('입찰 전략')
col_w3 = [20, 12, 12, 12, 14, 14, 14, 14, 14, 22]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = 1
merge(ws3, r, 1, r, 9, '26년 군인공제회 입찰 전략 분석', fill=H_FILL, font=Font(color='FFFFFF', bold=True, size=11))
r += 2

# 과거 입찰 히스토리
merge(ws3, r, 1, r, 9, '▶ 과거 입찰 결과 (나라장터 기준)', fill=T_FILL, font=T_FONT)
r += 1
for h, ci in zip(['연도', '1위', '입찰금액', '투찰률', '총할인율', '비고'], range(1,7)):
    hdr(ws3, r, ci, h)
r += 1

history = [
    ('24년', '에스티엠', '15.21억', '92.93%', '10.87%', '편의점+신세계3/5/7만원'),
    ('25년', '(에스티엠 추정)', '약 17.8억', '-', '11.18%', '편의점+신세계3/5/7만원'),
]
for row_data in history:
    for ci, v in enumerate(row_data, 1):
        cell_set(ws3, r, ci, v, fill=GRAY, font=NRM_FONT, align=C)
    r += 1

r += 1
merge(ws3, r, 1, r, 9, '▶ 26년 입찰 시나리오별 수익 요약 (긍정 미교환율 기준)', fill=T_FILL, font=T_FONT)
r += 1
for h, ci in zip(['할인율', '총거래금액', '총할인금액', '청구금액', '수수료수익', '미교환수익', '총수익', '수익율', '판단'], range(1,10)):
    hdr(ws3, r, ci, h)
r += 1

scenarios_pos = []
for disc in DISCOUNT_RATES:
    tot_거래 = tot_할인 = tot_수수료 = tot_미교환 = 0
    for name, 권종, 수수료율, 발송량, 미교환율_pos, _ in ITEMS:
        거래 = 권종 * 발송량
        할인 = round(거래 * disc)
        교환 = 1 - 미교환율_pos
        수수료 = round(거래 * 교환 * 수수료율)
        미교환 = round(거래 * 미교환율_pos)
        tot_거래   += 거래
        tot_할인   += 할인
        tot_수수료 += 수수료
        tot_미교환 += 미교환
    tot_청구  = tot_거래 - tot_할인
    tot_수익  = tot_미교환 + tot_수수료 - tot_할인
    tot_수익율 = tot_수익 / tot_거래
    판단 = '적극추천' if tot_수익율 >= 0.02 else ('검토' if tot_수익율 >= 0 else '손실위험')
    scenarios_pos.append((disc, tot_거래, tot_할인, tot_청구, tot_수수료, tot_미교환, tot_수익, tot_수익율, 판단))

for disc, 거래, 할인, 청구, 수수료, 미교환, 수익, 수익율, 판단 in scenarios_pos:
    fill = POS_FILL if 수익 > 0 else NEG_FILL
    fn   = GRN_FONT if 수익 > 0 else RED_FONT
    vals = [f'{disc*100:.0f}%', 거래, 할인, 청구, 수수료, 미교환, 수익, 수익율, 판단]
    fmts = [None,'#,##0','#,##0','#,##0','#,##0','#,##0','#,##0','0.00%',None]
    for ci, (v, f) in enumerate(zip(vals, fmts), 1):
        cell_set(ws3, r, ci, v, fill=fill, font=fn if ci in (7,8,9) else NRM_FONT, align=C, fmt=f)
    r += 1

r += 1
merge(ws3, r, 1, r, 9, '▶ 26년 vs 25년 변경사항 요약', fill=T_FILL, font=T_FONT)
r += 1
for h, ci in zip(['항목', '25년', '26년', '변화'], range(1,5)):
    hdr(ws3, r, ci, h)
r += 1
changes = [
    ('배정예산',    '20억원',     '23억원',     '+3억 (+15%)'),
    ('기초금액',    '약 19.67억', '약 22.87억', '+3.2억'),
    ('자격요건 실적', '15억원↑',  '20억원↑',   '+5억 상향'),
    ('편의점 수량', '17,431매',   '19,562매',   '+2,131매'),
    ('3만원 수량',  '7,208매',    '10,484매',   '+3,276매 (최대 증가)'),
    ('10만원~',     '없음',       '신규 추가',  '3만~35만 全권종'),
    ('40만원',      '없음',       '3매 신규',   '최고액권 추가'),
    ('경쟁사',      '에스티엠/즐거운/쿠프', '동일 예상', '25년 과 동일구도'),
]
for row_data in changes:
    alt = r % 2 == 0
    for ci, v in enumerate(row_data, 1):
        cell_set(ws3, r, ci, v, fill=GRAY if alt else None, font=NRM_FONT, align=L if ci == 1 else C)
    r += 1

r += 2
merge(ws3, r, 1, r, 9, '▶ 입찰 전략 제언', fill=T_FILL, font=T_FONT)
r += 1
tips = [
    '① 26년 총할인율 목표: 11~12% 수준 (24~25년 낙찰 흐름 기준)',
    '② 3만원 수량이 7,208→10,484으로 급증 → 낙전 수익 최대 기여 권종, 할인율 배분 시 집중',
    '③ 10만원 이상 신규 권종은 미교환율 데이터 부족 → 부정 시나리오(시트2) 수익도 반드시 확인',
    '④ 편의점 1만원 할인율을 높게 (20% 이상) 제시해 경쟁력 확보 → 낙전율 높아 실익 있음',
    '⑤ 총 낙찰가 목표: 약 20~21억원 (청구금액 기준)',
]
for tip in tips:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    cell = ws3.cell(r, 1, tip)
    cell.font = Font(size=9)
    cell.alignment = L
    cell.fill = NOTE_FILL
    cell.border = BORDER
    r += 1

ws3.freeze_panes = 'A4'

# ══════════════════════════════════════════════════════════
# 시트 4: 과거 입찰데이터 (원본 복사)
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet('25년 시트별 인사이트')
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 70
r = 1
merge(ws4, r, 1, r, 2, '25년 시뮬레이션 파일 시트별 인사이트', fill=H_FILL, font=Font(color='FFFFFF', bold=True, size=11))
r += 1

insights = [
    ('시트', '인사이트'),
    ('확정 할인율',
     '■ 7.4% 할인율: 예상수익 -3.6% (약 -5,231만원 손실)\n'
     '■ 6.03% 할인율: 예상수익 -1.04% (약 -2,042만원 손실)\n'
     '→ 25년 시뮬레이션 당시 어떤 할인율로 잡아도 수익 플러스가 어려웠음\n'
     '→ 24년 낙찰사(에스티엠) 총할인율 10.87% 기준으로 맞추면 미교환율이 충분히 높아야 수익\n'
     '→ 실제 운영 결과(로우데이터): 3만원 avg 50.3%, 합산 낙전수익 8개월 2.69억 → 예측보다 좋았음'),
    ('시뮬레이션',
     '■ 긍정: 할인율 7%, 편의점 30%/3만원 7% 미교환율 → 수익 3,823만원 (수익률 1.94%)\n'
     '■ 부정: 할인율 7%, 보수적 미교환율 → 손실 -2,412만원 (-1.23%)\n'
     '■ 마케팅 추가 (버거킹/커피/치킨 연 6회): 수익 4,000만원 (수익률 8.5%) 별도 기대\n'
     '→ 회원복지 손익분기를 마케팅 수익으로 커버하는 포트폴리오 전략을 검토했었음'),
    ('시뮬레이션 (2)',
     '■ 권종별 차등 할인율 적용 시나리오 (편의점 27.18%, 백화점 고액 2.54%)\n'
     '■ 낙전 높은 편의점에 높은 할인율 → 경쟁력 확보 + 낙전으로 수익 보전 전략\n'
     '■ 5만원 수량 5,164→12,449 조정 시나리오: 수익 1,970만원 (1.5%) → 수량 배분이 핵심\n'
     '→ 26년에도 동일 로직 적용 가능: 3만원 10,484매 대폭 증가로 낙전 수익 구조 개선'),
    ('과거 입찰데이터',
     '■ 24년: 에스티엠 1위 (15.21억, 투찰률 92.93%), 즐거운 2위 (15.35억), 쿠프 3위 (15.51억)\n'
     '■ 낙찰 총할인율: 에스티엠 10.87%, 즐거운 5.5%, 쿠프 4.5%\n'
     '■ 에스티엠 실제 단가: 편의점 CU 20% 할인 / 신세계 3만 7% / 5만 4% / 7만 3.8%\n'
     '■ 25년 추정 총할인율: 11.18% (예산 20억, 기초금액 19.67억)\n'
     '→ 26년 예산 23억 / 기초금액 약 22.87억 기준, 경쟁 입찰 시 총할인율 11~12% 예상'),
    ('하이닉스',
     '■ 이천: 네이버페이5만(3.65%) + 신세계5만(6.84%) → 전체 수익률 4.44%\n'
     '■ 청주: 네이버페이 25/50만원 → 수익률 7.53%\n'
     '→ 할인율 1~1.5%로 낮고 미교환율 3~7%인 구조 → 수익률 낮음\n'
     '→ 군인공제회 대비 낙전 기대치 낮음 (단기/복지 이벤트 성격)'),
    ('생일24',
     '■ 신세계 5만원 62,504매 발송 / 사용률 93.2% / 미교환율 6.8%\n'
     '■ 수익률 3.59% (수익금 1.12억)\n'
     '→ 대규모 B2C성 발송은 미교환율이 낮음 → 군인공제회 미교환율과 비교 시 군인공제회가 훨씬 유리'),
    ('우리카드',
     '■ 소량(1~344매) 단건 이벤트 → 미교환율 0~100% 편차 큼\n'
     '■ 할인율 3.5~5.5% 수준 / 1개 교환되면 수수료가 비용 초과 → 손실 발생 다수\n'
     '→ 소량 B2B는 낙전보다 수수료 수익에 의존해야 하는 구조\n'
     '→ 군인공제회 같은 대규모 건과 수익 모델이 다름'),
]
for row_data in insights:
    for ci, v in enumerate(row_data, 1):
        is_hdr = (row_data[0] == '시트')
        fill = H_FILL if is_hdr else (GRAY if r % 2 == 0 else None)
        fn   = H_FONT if is_hdr else NRM_FONT
        cell_set(ws4, r, ci, v, fill=fill, font=fn, align=C if ci == 1 else L)
        ws4.row_dimensions[r].height = 80 if not is_hdr else 16
    r += 1

ws4.freeze_panes = 'A3'

# 첫 번째 기본 시트 삭제
del wb['Sheet']

OUTPUT = r'c:\Users\jihye\tutorial\군인공제회\군인공제회 26년 시뮬레이션 초안.xlsx'
wb.save(OUTPUT)
print(f'저장 완료: {OUTPUT}')
print(f'시트 목록: {wb.sheetnames}')
