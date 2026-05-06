import os, re, sys
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\jihye\tutorial\군인공제회\202508-202603 로우데이터'
output = r'c:\Users\jihye\tutorial\군인공제회\202508-202603_병합.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '전체병합'

header_written = False
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(color='FFFFFF', bold=True)
alt_fill = PatternFill('solid', fgColor='DCE6F1')

row_num = 1

files = sorted([f for f in os.listdir(folder) if f.endswith('.xls')])

for fname in files:
    # 파일명에서 월 추출 (20250801~20250831 → 2025-08)
    m = re.search(r'(\d{4})(\d{2})\d{2}~', fname)
    month_label = f"{m.group(1)}-{m.group(2)}" if m else fname

    fpath = os.path.join(folder, fname)
    with open(fpath, encoding='euc-kr', errors='replace') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    rows = soup.find_all('tr')
    if not rows:
        continue

    for i, row in enumerate(rows):
        cells = [td.get_text(strip=True) for td in row.find_all('td')]
        if not cells:
            continue

        if i == 0:  # 헤더 행
            if not header_written:
                full_row = ['월'] + cells
                ws.append(full_row)
                for cell in ws[row_num]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                row_num += 1
                header_written = True
        else:
            full_row = [month_label] + cells
            ws.append(full_row)
            if row_num % 2 == 0:
                for cell in ws[row_num]:
                    cell.fill = alt_fill
            row_num += 1

# 열 너비 자동 조정
for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

wb.save(output)
print(f'완료: {output}')
print(f'총 데이터 행 수: {row_num - 2}')
